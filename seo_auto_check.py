from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import subprocess
from datetime import datetime, date
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import calendar
import json
import logging

# =============================================================================
# 設定常數
# =============================================================================
INPUT_FILE = "data/input/seo_search_keyword.xlsx"
OUTPUT_DIR = "data/output"
CONVERTED_DIR = "data/input/converted"
LOG_DIR = "logs"

# =============================================================================
# 輔助函式
# =============================================================================

def setup_logger():
    """設定 logger，同時輸出到檔案和終端"""
    os.makedirs(LOG_DIR, exist_ok=True)
    log_filename = f"seo_{date.today().strftime('%Y-%m-%d')}.log"
    log_path = os.path.join(LOG_DIR, log_filename)

    logger = logging.getLogger("seo_auto_check")
    logger.setLevel(logging.DEBUG)

    # 避免重複加入 handler（多次呼叫時）
    if logger.handlers:
        return logger

    formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s',
                                  datefmt='%Y-%m-%d %H:%M:%S')

    file_handler = logging.FileHandler(log_path, mode='a', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)

    return logger


def _convert_oragain_to_final(oragain_path, final_path):
    """將 oragain 格式轉換為 final 格式並存檔"""
    logger = logging.getLogger("seo_auto_check")
    with open(oragain_path, 'r', encoding='utf-8') as f:
        raw_data = json.load(f)

    converted = []
    for entry in raw_data:
        converted.append({
            "date": entry["西元日期"],
            "week": entry["星期"],
            "isHoliday": entry["是否放假"] == "2",
            "description": entry.get("備註", "")
        })

    os.makedirs(os.path.dirname(final_path), exist_ok=True)
    with open(final_path, 'w', encoding='utf-8') as f:
        json.dump(converted, f, ensure_ascii=False, indent=2)

    logger.info(f"已從 {oragain_path} 轉換建立 {final_path}（{len(converted)} 筆）")


def load_holidays(year):
    """從 JSON 檔讀取假日資料，回傳 set of date，失敗回傳 None"""
    logger = logging.getLogger("seo_auto_check")

    final_path = os.path.join(CONVERTED_DIR, "final", f"converted_{year}.json")
    oragain_path = os.path.join(CONVERTED_DIR, "oragain", f"{year}_export.json")

    # final 不存在 → 嘗試從 oragain 轉換建立
    if not os.path.exists(final_path):
        if os.path.exists(oragain_path):
            try:
                _convert_oragain_to_final(oragain_path, final_path)
            except Exception as e:
                logger.error(f"轉換 {oragain_path} 失敗: {e}")
                return None
        else:
            logger.error(f"找不到 {year} 年的假日資料檔（已搜尋 final/ 和 oragain/）")
            return None

    # 讀取 final
    try:
        with open(final_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        holidays = set()
        for entry in data:
            if entry.get('isHoliday') is True:
                d = datetime.strptime(entry['date'], '%Y%m%d').date()
                holidays.add(d)
        logger.info(f"從 {final_path} 載入 {len(holidays)} 個假日")
        return holidays
    except Exception as e:
        logger.error(f"讀取 {final_path} 失敗: {e}")
        return None


def load_config():
    """從 CONFIG 工作表讀取設定"""
    logger = logging.getLogger("seo_auto_check")
    if not os.path.exists(INPUT_FILE):
        logger.error(f"找不到設定檔 {INPUT_FILE}")
        return None

    wb = load_workbook(INPUT_FILE, data_only=True)
    if 'CONFIG' not in wb.sheetnames:
        logger.error("找不到 CONFIG 工作表")
        return None

    ws = wb['CONFIG']
    config = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=False):
        key_cell, val_cell = row[0], row[1]
        if key_cell.value:
            config[str(key_cell.value).strip()] = val_cell.value

    # 預設值
    config.setdefault('target_domain', '')
    config.setdefault('search_depth_pages', 2)
    config.setdefault('outside_threshold', 15)
    config.setdefault('output_filename', 'seo_search_results.xlsx')

    # 確保數值型別
    config['search_depth_pages'] = int(config['search_depth_pages'])
    config['outside_threshold'] = int(config['outside_threshold'])

    return config


def load_keywords():
    """從 KEYWORDS 工作表讀取關鍵字 + 定價資料"""
    logger = logging.getLogger("seo_auto_check")
    if not os.path.exists(INPUT_FILE):
        logger.error(f"找不到關鍵字檔案 {INPUT_FILE}")
        return []

    wb = load_workbook(INPUT_FILE, data_only=True)
    if 'KEYWORDS' not in wb.sheetnames:
        logger.error("找不到 KEYWORDS 工作表")
        return []

    ws = wb['KEYWORDS']
    keywords_data = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True), start=2):
        keyword = row[0]
        if keyword is not None and isinstance(keyword, (int, float)):
            logger.warning(f"  KEYWORDS 第 {row_idx} 列的 B 欄為數字 ({keyword})，已跳過")
        elif keyword and str(keyword).strip():
            entry = {
                'keyword': str(keyword).strip(),
                'tier1_condition': int(row[1]) if row[1] else 10,
                'tier1_price': int(row[2]) if row[2] else 0,
                'tier2_condition': int(row[3]) if row[3] else None,
                'tier2_price': int(row[4]) if row[4] else None,
            }
            keywords_data.append(entry)

    logger.info(f"成功載入 {len(keywords_data)} 個關鍵字")
    return keywords_data


def is_holiday(check_date, holidays_set):
    """判斷是否為假日（依據 JSON 假日資料）"""
    return check_date in holidays_set


def format_ranking(rank, page, threshold):
    """格式化排名為顯示字串"""
    if rank is None:
        return f"{threshold}以外"
    return f"{rank}/P.{page}"


def search_keyword(driver, keyword, target_domain, max_pages):
    """搜尋關鍵字，在多頁結果中尋找目標網域的排名位置"""
    logger = logging.getLogger("seo_auto_check")
    try:
        driver.get("https://www.google.com")
        time.sleep(2)

        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()

        # 模擬人類打字速度
        for char in keyword:
            search_box.send_keys(char)

        search_box.send_keys(Keys.RETURN)
        time.sleep(5)

        # 將 target_domain 正規化：去掉 scheme / www. 前綴，取得純域名
        # 例如 "www.scincotaiwan.tw" → "scincotaiwan.tw"
        # 例如 "https://www.scincotaiwan.tw" → "scincotaiwan.tw"
        parsed_target = urlparse(target_domain)
        target_host = parsed_target.netloc or parsed_target.path or target_domain
        target_host = target_host.lower().strip("/")
        if target_host.startswith("www."):
            target_base = target_host[4:]
        else:
            target_base = target_host
        logger.info(f"  比對目標域名: {target_base} (原始: {target_domain})")
        cumulative_rank = 0

        for page_num in range(1, max_pages + 1):
            # 用標題連結 (含 h3 的 a 標籤) 來抓每筆搜尋結果的 URL
            title_links = driver.find_elements(By.CSS_SELECTOR, "#search a:has(h3)")
            # 備援：若新版 selector 抓不到，退回舊版
            if not title_links:
                title_links = driver.find_elements(By.CSS_SELECTOR, "#search .g a[href]")
            logger.info(f"  第 {page_num} 頁：找到 {len(title_links)} 個結果標題連結")

            for link in title_links:
                try:
                    href = link.get_attribute("href")
                    if not href:
                        logger.info(f"    [跳過] 無 href")
                        continue

                    parsed = urlparse(href)
                    result_host = parsed.netloc.lower()

                    # 排除 Google 自身的連結
                    if 'google.com' in result_host:
                        logger.info(f"    [跳過] Google 連結: {result_host}")
                        continue

                    cumulative_rank += 1

                    # 去掉 www. 後比對，只要結果域名「包含」目標域名即可
                    # 例如 target_base="scincotaiwan.tw"
                    # 會匹配: scincotaiwan.tw, www.scincotaiwan.tw, shop.scincotaiwan.tw
                    result_base = result_host[4:] if result_host.startswith("www.") else result_host
                    matched = target_base in result_base

                    logger.info(f"    #{cumulative_rank} {result_host} → {'✓ 匹配' if matched else '✗'}")

                    if matched:
                        logger.info(f"  找到！排名第 {cumulative_rank} 名，第 {page_num} 頁")
                        return {"rank": cumulative_rank, "page": page_num}

                except Exception as e:
                    logger.info(f"    [跳過] 解析結果時發生錯誤: {e}")
                    continue

            # 嘗試翻到下一頁
            if page_num < max_pages:
                try:
                    next_btn = driver.find_element(By.ID, "pnnext")
                    next_btn.click()
                    time.sleep(3)
                except Exception:
                    logger.warning(f"  無法翻到第 {page_num + 1} 頁")
                    break

        logger.info(f"  未找到目標網域（已搜尋 {max_pages} 頁）")
        return None

    except Exception as e:
        logger.error(f"  搜尋關鍵字「{keyword}」時發生錯誤: {e}")
        return None


# =============================================================================
# Excel 輸出函式
# =============================================================================

def init_month_sheet(wb, year_month, keywords_data, config):
    """建立或取得月份工作表，寫入列 1-7 結構，預填日期與假日"""
    sheet_name = year_month.strftime("%Y-%m")

    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    num_kw = len(keywords_data)
    holidays_set = config['holidays_set']

    # --- 列 1-3: Tier2 定價（只有有 tier2 的關鍵字才填） ---
    ws.cell(row=1, column=1, value="條件(名次)")
    ws.cell(row=2, column=1, value="價格(未稅)")
    ws.cell(row=3, column=1, value="價格(含稅)")

    for i, kw in enumerate(keywords_data):
        col = i + 2  # B=2, C=3, ...
        if kw['tier2_condition'] is not None:
            ws.cell(row=1, column=col, value=kw['tier2_condition'])
            ws.cell(row=2, column=col, value=kw['tier2_price'])
            # 含稅用公式
            col_letter = get_column_letter(col)
            ws.cell(row=3, column=col, value=f"={col_letter}2*1.05")

    # S 欄警示文字
    s_col = num_kw + 2  # 關鍵字後一欄
    ws.cell(row=1, column=s_col, value="**未超過11個關鍵字不收費**")
    ws.cell(row=1, column=s_col).font = Font(name='Arial', size=10, color='FFFF0000')

    # --- 列 4-6: Tier1 定價 ---
    ws.cell(row=4, column=1, value="條件(名次)")
    ws.cell(row=5, column=1, value="價格(未稅)")
    ws.cell(row=6, column=1, value="價格(含稅)")

    for i, kw in enumerate(keywords_data):
        col = i + 2
        ws.cell(row=4, column=col, value=kw['tier1_condition'])
        ws.cell(row=5, column=col, value=kw['tier1_price'])
        col_letter = get_column_letter(col)
        ws.cell(row=6, column=col, value=f"={col_letter}5*1.05")

    # --- 列 7: 關鍵字標題列 ---
    ws.cell(row=7, column=1, value="關鍵字")
    for i, kw in enumerate(keywords_data):
        col = i + 2
        ws.cell(row=7, column=col, value=kw['keyword'])

    # --- 列 8+: 預填當月所有日期 ---
    year = year_month.year
    month = year_month.month
    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        row_num = 7 + day  # 第 8 列 = 第 1 天
        current_date = date(year, month, day)
        ws.cell(row=row_num, column=1, value=current_date)
        ws.cell(row=row_num, column=1).number_format = 'mm-dd-yy'

        # 假日標記
        if is_holiday(current_date, holidays_set):
            for col in range(2, num_kw + 2):
                ws.cell(row=row_num, column=col, value="休")

    return ws


def save_ranking(ws, target_date, col_index, ranking_str, check_time=None):
    """寫入單筆排名到對應的日期列和關鍵字欄"""
    # 找到日期對應的列
    day = target_date.day
    row_num = 7 + day  # 第 8 列 = 第 1 天

    ws.cell(row=row_num, column=col_index, value=ranking_str)

    # 寫入時間到最後一欄（如果有提供）
    if check_time is not None:
        # 找到時間欄（關鍵字欄後一欄）
        time_col = ws.max_column
        # 搜尋已存在的時間欄位，或使用關鍵字數+2
        for c in range(2, ws.max_column + 2):
            cell_val = ws.cell(row=1, column=c).value
            if cell_val and '未超過' in str(cell_val):
                time_col = c
                break
        ws.cell(row=row_num, column=time_col, value=check_time)
        ws.cell(row=row_num, column=time_col).number_format = 'h:mm'


def format_sheet(ws, num_keywords):
    """套用所有格式（框線、填充、對齊、凍結、欄寬）"""
    # 樣式定義
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    cyan_fill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    default_font = Font(name='Arial', size=10)
    price_format = '#,##0'

    last_kw_col = num_keywords + 1  # 最後一個關鍵字欄
    time_col = num_keywords + 2     # 時間欄

    # --- 欄寬 ---
    ws.column_dimensions['A'].width = 12
    for i in range(2, last_kw_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(time_col)].width = 25

    # --- 凍結窗格 ---
    ws.freeze_panes = 'B8'

    # --- 套用格式到所有有資料的儲存格 ---
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=time_col):
        for cell in row:
            cell.font = default_font
            cell.alignment = center_align

    # --- 列 1-7 框線 ---
    for row_num in range(1, 8):
        for col in range(1, last_kw_col + 1):
            ws.cell(row=row_num, column=col).border = thin_border

    # --- 價格數字格式 ---
    for col in range(2, last_kw_col + 1):
        ws.cell(row=2, column=col).number_format = price_format
        ws.cell(row=3, column=col).number_format = price_format
        ws.cell(row=5, column=col).number_format = price_format
        ws.cell(row=6, column=col).number_format = price_format

    # --- 含稅列青色填充（列 3 和列 6） ---
    for col in range(2, last_kw_col + 1):
        ws.cell(row=3, column=col).fill = cyan_fill
        ws.cell(row=6, column=col).fill = cyan_fill

    # --- 有 tier2 的關鍵字在列 7 標黃色 ---
    for col in range(2, last_kw_col + 1):
        # 如果列 1 有值（tier2 condition），則該關鍵字有 tier2
        if ws.cell(row=1, column=col).value is not None:
            ws.cell(row=7, column=col).fill = yellow_fill

    # --- S1 警示文字紅色字型 ---
    ws.cell(row=1, column=time_col).font = Font(
        name='Arial', size=10, color='FFFF0000'
    )

    # --- 資料列框線 ---
    for row_num in range(8, ws.max_row + 1):
        for col in range(1, time_col + 1):
            ws.cell(row=row_num, column=col).border = thin_border


# =============================================================================
# 主程式
# =============================================================================

def main():
    """主程式"""
    logger = setup_logger()

    logger.info("=" * 60)
    logger.info("  SEO 自動排名追蹤工具")
    logger.info("  目標網域排名監測 — Pivot 矩陣輸出")
    logger.info("=" * 60)

    # 1. 載入設定
    config = load_config()
    if not config:
        return

    # 2. 載入關鍵字
    keywords_data = load_keywords()
    if not keywords_data:
        logger.error("沒有找到關鍵字，程式結束。")
        return

    # 3. 驗證 target_domain
    target_domain = str(config['target_domain']).strip()
    if not target_domain:
        logger.error("CONFIG 工作表中的 target_domain 未設定！")
        logger.error("請在 data/input/seo_search_keyword.xlsx 的 CONFIG 工作表中填入目標網域。")
        return

    # 4. 載入假日資料
    today = date.today()
    holidays_set = load_holidays(today.year)
    if holidays_set is None:
        logger.error("無法載入假日資料，程式停止。")
        return
    config['holidays_set'] = holidays_set

    logger.info(f"目標網域: {target_domain}")
    logger.info(f"搜尋深度: {config['search_depth_pages']} 頁")
    logger.info(f"超出閾值: {config['outside_threshold']}")
    logger.info(f"關鍵字數: {len(keywords_data)}")

    # 5. 開啟/建立輸出 Excel
    output_file = os.path.join(OUTPUT_DIR, config['output_filename'])
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
    else:
        wb = Workbook()
        # 移除預設工作表
        wb.remove(wb.active)

    # 6. 建立當月工作表
    year_month = date(today.year, today.month, 1)
    ws = init_month_sheet(wb, year_month, keywords_data, config)

    # 7. 假日檢查
    if is_holiday(today, config['holidays_set']):
        logger.info(f"今天 ({today}) 為假日，已標記為「休」。")
        format_sheet(ws, len(keywords_data))
        wb.save(output_file)
        logger.info(f"已儲存至 {output_file}")
        return

    # 8. 檢查今天是否已有資料（避免重複覆蓋）
    day_row = 7 + today.day
    existing_val = ws.cell(row=day_row, column=2).value
    if existing_val and existing_val != "休":
        logger.info(f"今天 ({today}) 已有資料（{existing_val}），跳過以避免覆蓋。")
        logger.info("如需重新執行，請先手動清除當日資料。")
        return

    # 9. 啟動 Chrome
    logger.info("正在啟動 Chrome 瀏覽器...")
    chrome_options = Options()
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)

    try:
        # 10. 逐一搜尋關鍵字
        for i, kw_data in enumerate(keywords_data):
            keyword = kw_data['keyword']
            col_index = i + 2  # B=2, C=3, ...

            logger.info(f"[{i+1}/{len(keywords_data)}] 搜尋關鍵字：{keyword}")

            result = search_keyword(
                driver, keyword, target_domain,
                config['search_depth_pages']
            )

            # 格式化排名
            if result:
                ranking_str = format_ranking(
                    result['rank'], result['page'],
                    config['outside_threshold']
                )
            else:
                ranking_str = format_ranking(
                    None, None,
                    config['outside_threshold']
                )

            logger.info(f"  結果: {ranking_str}")

            # 寫入排名
            check_time = datetime.now().time()
            save_ranking(ws, today, col_index, ranking_str,
                         check_time if i == len(keywords_data) - 1 else None)

            # 每個關鍵字後儲存檔案（防當機遺失）
            wb.save(output_file)

            # 關鍵字間等待
            if i < len(keywords_data) - 1:
                logger.info("  等待 5 秒...")
                time.sleep(5)

        # 11. 套用最終格式
        format_sheet(ws, len(keywords_data))

        # 12. 最終儲存
        wb.save(output_file)
        logger.info("=" * 60)
        logger.info(f"所有關鍵字搜尋完成！共處理 {len(keywords_data)} 個關鍵字")
        logger.info(f"結果已儲存至 {output_file}")
        logger.info("=" * 60)

    finally:
        # 關閉 Chrome（macOS AppleScript 方式）
        try:
            applescript = '''
            tell application "Google Chrome"
                if (count of windows) > 0 then
                    close (window 1)
                end if
            end tell
            '''
            subprocess.run(['osascript', '-e', applescript], check=True, timeout=5)
            time.sleep(1)
        except Exception as e:
            logger.warning(f"AppleScript 關閉視窗失敗: {e}")

        try:
            driver.quit()
        except Exception as e:
            logger.warning(f"driver.quit() 失敗: {e}")

        try:
            service.stop()
        except Exception as e:
            logger.warning(f"service.stop() 失敗: {e}")


if __name__ == "__main__":
    main()
